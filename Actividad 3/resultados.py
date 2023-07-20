import stanza
from tabulate import tabulate
from openpyxl import Workbook

ra1 = "Elige y evalúa los factores influyentes que permitan definir de forma apropiada la organización física requerida para un implementar una base de datos"
ra2 = "Propone soluciones innovadoras a problemas en la sociedad sistemáticos y metodológicos del pensamiento de diseño."
ra3 = "Aplica la metodología del pensamiento de diseño para identificar problemas locales relacionados a la ingeniería en computación."
ra4 = "Diseña prototipos aplicando técnicas y herramientas de prototipado rápido de Ingeniería en Computación para crear soluciones innovadoras"
ra5 = "Diseña prototipos aplicando técnicas y herramientas de prototipado rápido de Ingeniería en Computación para crear soluciones innovadoras"

# Inicializar el modelo de procesamiento de idioma español
nlp = stanza.Pipeline('es')

# Textos de ejemplo
textos = [ra1, ra2, ra3, ra4, ra5]

# Lista para almacenar los resultados
resultados = []

# Procesar cada texto en la lista
for i, texto in enumerate(textos, start=1):
    # Procesar el texto
    doc = nlp(texto)


    # Contadores para cada tipo de palabra
    contador_pronombres = 0
    contador_sustantivos = 0
    contador_adjetivos = 0
    contador_verbos = 0
    contador_adverbios = 0

    # Contadores para cada tipo de verbo
    contador_auxiliares = 0
    contador_principales = 0
    contador_tiempos = 0
    contador_infinitivo = 0


    
    # Recorrer cada oración en el documento
    for sentence in doc.sentences:
        # Recorrer cada palabra en la oración
        for word in sentence.words:
            # Contar tipos de palabras
            if word.upos == 'PRON':
                contador_pronombres += 1
            elif word.upos == 'NOUN':
                contador_sustantivos += 1
            elif word.upos == 'ADJ':
                contador_adjetivos += 1
            elif word.upos == 'ADV':
                contador_adverbios += 1
            elif word.upos == 'AUX':
                contador_auxiliares += 1
            elif word.upos == 'VERB':
                contador_verbos += 1
                if word.deprel == 'root':
                    contador_principales += 1
                if word.feats != "Tense=Pres" and word.feats != 'VerbForm=Inf':
                    contador_tiempos += 1
                if word.feats == 'VerbForm=Inf':
                    contador_infinitivo += 1
            
    # Guardar los resultados en una lista
    resultados.append([f"RA{i}", contador_pronombres, contador_sustantivos, contador_adjetivos, contador_verbos,
                       contador_adverbios, contador_auxiliares, contador_principales, contador_tiempos,
                       contador_infinitivo])
    


# Escribir los resultados en la hoja de cálculo
headers = ['RA', 'Pronombres', 'Sustantivos', 'Adjetivos', 'Verbos', 'Adverbios', 'Auxiliares', 'Principales',
           'Otros tiempos', 'Infinitivo']

# Imprimir los resultados en forma de tabla
print(tabulate(resultados, headers=headers))



# Crear un nuevo archivo de Excel
wb = Workbook()

# Crear una hoja de cálculo
sheet = wb.active
sheet.title = "Resultados"

# Escribir los encabezados
for col_num, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header

# Escribir los resultados
for row_num, resultado in enumerate(resultados, start=2):
    for col_num, value in enumerate(resultado, start=1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.value = value

# Guardar el archivo de Excel
wb.save("resultados.xlsx")